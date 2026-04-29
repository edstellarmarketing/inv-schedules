from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
import json
import re as _re

from data import DAY_TO_US_DOW, COURSE_PRICING

_INPUT_TZ = "America/New_York"

_VALID_DAYS = set(DAY_TO_US_DOW.keys())

# Common abbreviation for each IANA timezone used in this app
_TZ_ABBR_MAP: dict[str, str] = {
    "America/New_York":    "EST",
    "America/Toronto":     "EST",
    "Europe/London":       "GMT",
    "Europe/Berlin":       "CET",
    "Europe/Amsterdam":    "CET",
    "America/Sao_Paulo":   "BRT",
    "Pacific/Auckland":    "NZST",
    "Asia/Dubai":          "GST",
    "Asia/Singapore":      "SGT",
    "Asia/Riyadh":         "AST",
    "Asia/Qatar":          "AST",
    "Asia/Kuala_Lumpur":   "MYT",
    "Asia/Tokyo":          "JST",
    "Asia/Colombo":        "SLST",
    "Asia/Dhaka":          "BST",
    "Africa/Johannesburg": "SAST",
    "Africa/Nairobi":      "EAT",
    "Africa/Lagos":        "WAT",
    "Asia/Kolkata":        "IST",
    "Australia/Sydney":    "AEST",
}


def _tz_abbr(tz_name: str) -> str:
    """Return a short, human-readable abbreviation for an IANA timezone name."""
    if tz_name in _TZ_ABBR_MAP:
        return _TZ_ABBR_MAP[tz_name]
    try:
        # Use a fixed winter date so we always get the standard (non-DST) name
        dt = datetime(2026, 1, 15, 12, 0, tzinfo=ZoneInfo(tz_name))
        abbr = dt.strftime("%Z")
        if not abbr.startswith(("+", "-")):
            return abbr
    except Exception:
        pass
    return tz_name


def to_12h(time_24: str) -> str:
    """Convert 'HH:MM' 24-hour string to 'H:MM AM/PM'."""
    h, m = map(int, time_24.split(":"))
    suffix = "AM" if h < 12 else "PM"
    h12 = h % 12 or 12
    return f"{h12}:{m:02d} {suffix}"


def _time_category(time_24: str) -> str:
    """
    Return the display time-slot category for a local 24-hour start time.
    Covers all 24 hours so every record gets a value.

    00:00 – 05:59  Early Morning
    06:00 – 08:59  Early Morning
    09:00 – 11:59  Morning
    12:00 – 14:59  Noon
    15:00 – 16:29  Early Evening
    16:30 – 17:59  Evening
    18:00 – 23:59  Late Evening
    """
    h, m = map(int, time_24.split(":"))
    mins = h * 60 + m
    if mins < 9*60:                return "Early Morning"   # midnight–9 AM
    if mins < 12*60:               return "Morning"         # 9 AM–noon
    if mins < 15*60:               return "Noon"            # noon–3 PM
    if mins < 16*60 + 30:          return "Early Evening"   # 3 PM–4:30 PM
    if mins < 18*60:               return "Evening"         # 4:30 PM–6 PM
    return "Late Evening"                                   # 6 PM–midnight


# Per-timezone minute offset applied on top of the standard DST-aware conversion.
# Positive = add minutes, negative = subtract minutes.
_TZ_OFFSET_ADJUST: dict[str, int] = {
    "Asia/Kolkata": -30,   # India: display 30 min earlier than raw conversion
}


def _convert_time(time_str: str, to_tz: str, ref_date: date) -> str:
    """
    Convert HH:MM from America/New_York to `to_tz` using `ref_date` for DST accuracy.
    Applies any entry in _TZ_OFFSET_ADJUST after conversion.
    Returns the original string unchanged when target is already New York.
    """
    if to_tz == _INPUT_TZ:
        return time_str
    h, m = map(int, time_str.split(":"))
    dt = datetime(ref_date.year, ref_date.month, ref_date.day, h, m,
                  tzinfo=ZoneInfo(_INPUT_TZ))
    result = dt.astimezone(ZoneInfo(to_tz)).strftime("%H:%M")
    adjust = _TZ_OFFSET_ADJUST.get(to_tz, 0)
    if adjust:
        rh, rm = map(int, result.split(":"))
        total = (rh * 60 + rm + adjust) % (24 * 60)
        result = f"{total // 60:02d}:{total % 60:02d}"
    return result


def _nth_day_of_month(year: int, month: int, anchor_us_dow: int, n: int) -> date | None:
    """
    Return the n-th occurrence of anchor_us_dow (US convention: Sun=0 … Sat=6)
    in the given month, or None if the month has fewer than n occurrences.
    """
    first_day = date(year, month, 1)
    first_us_dow = (first_day.weekday() + 1) % 7          # Python Mon=0 → US Sun=0
    days_ahead = (anchor_us_dow - first_us_dow) % 7
    result = first_day + timedelta(days=days_ahead) + timedelta(weeks=n - 1)
    return result if result.month == month else None


def _batch_dates_for_type(
    batch_type: str,
    day_format: list[str],
    weeks: list[int],
    from_date: date,
    to_date: date,
    training_days: int,
) -> list[tuple[date, date, list[date]]]:
    """
    Return list of (start, end, sessions) tuples for the given batch type.

    Anchor logic:
    - weekday: W_n = n-th Monday of the month; batch starts on that Monday.
    - weekend: W_n = n-th Saturday of the month; batch starts on the first
      day of day_format relative to that Saturday (e.g. Friday before it
      when format begins on Fri).

    A batch is skipped when its first session falls outside the current
    calendar month or outside [from_date, to_date].
    """
    if not day_format or not weeks:
        return []

    first_us_dow = DAY_TO_US_DOW[day_format[0]]

    if batch_type == "weekday":
        anchor_us_dow = DAY_TO_US_DOW["Mon"]   # always anchor on Monday
        start_offset = 0                        # batch starts on the Monday itself
    else:
        anchor_us_dow = DAY_TO_US_DOW["Sat"]   # always anchor on Saturday
        start_offset = first_us_dow - anchor_us_dow  # e.g. Fri(5)-Sat(6) = -1

    batches: list[tuple[date, date, list[date]]] = []

    current = from_date.replace(day=1)
    while current <= to_date.replace(day=1):
        year, month = current.year, current.month

        for week_num in sorted(weeks):
            anchor_date = _nth_day_of_month(year, month, anchor_us_dow, week_num)
            if anchor_date is None:
                continue

            first_session = anchor_date + timedelta(days=start_offset)

            # Skip if first session fell into the previous month
            if first_session.month != month:
                continue
            if first_session < from_date:
                continue

            sessions = [first_session + timedelta(days=i) for i in range(training_days)]

            if sessions[-1] > to_date:
                continue

            batches.append((sessions[0], sessions[-1], sessions))

        current = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    return batches


def _calc_final_price(base_price: int, tier_pct: int, exchange_rate) -> float | None:
    if exchange_rate is None:
        return None
    return round(base_price * (1 + tier_pct / 100) * exchange_rate, 2)


def _next_occurrence_after(base: date, target_us_dow: int) -> date:
    """
    Returns earliest date strictly AFTER `base` with the given US day-of-week (Sun=0..Sat=6).
    """
    base_us_dow = (base.weekday() + 1) % 7   # Python Mon=0 → US Sun=0
    delta = (target_us_dow - base_us_dow) % 7
    if delta == 0:
        delta = 7
    return base + timedelta(days=delta)


def _sessions_from_sequence(anchor: date, day_sequence: list[str]) -> list[date]:
    """
    Builds session dates by walking day_sequence forward from anchor.
    sessions[0] = anchor
    sessions[i] = _next_occurrence_after(sessions[i-1], DAY_TO_US_DOW[day_sequence[i]])
    """
    if not day_sequence:
        return []
    sessions = [anchor]
    for i in range(1, len(day_sequence)):
        nxt = _next_occurrence_after(sessions[i - 1], DAY_TO_US_DOW[day_sequence[i]])
        sessions.append(nxt)
    return sessions


def _batch_dates_from_sequence(
    anchor_us_dow: int,
    day_sequence: list[str],
    weeks: list[int],
    from_date: date,
    to_date: date,
) -> list[tuple[date, date, list[date]]]:
    """
    Generic batch generator using arbitrary anchor day and day sequence.
    Iterates months in [from_date, to_date], finds nth occurrence of anchor_us_dow,
    generates sessions via _sessions_from_sequence, skips invalid batches.
    """
    if not day_sequence or not weeks:
        return []

    batches: list[tuple[date, date, list[date]]] = []

    current = from_date.replace(day=1)
    while current <= to_date.replace(day=1):
        year, month = current.year, current.month

        for week_num in sorted(weeks):
            anchor = _nth_day_of_month(year, month, anchor_us_dow, week_num)
            if anchor is None:
                continue
            if anchor < from_date:
                continue
            if anchor.month != month:
                continue

            sessions = _sessions_from_sequence(anchor, day_sequence)
            if not sessions:
                continue

            if sessions[-1] > to_date:
                continue

            batches.append((sessions[0], sessions[-1], sessions))

        current = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    return batches


def _parse_time_12h(s: str) -> str:
    """
    Parses '9:00 AM' → '09:00', '5:00 PM' → '17:00'.
    Tries formats: "%I:%M %p", "%I:%M%p", "%H:%M".
    """
    s = str(s).strip()
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).strftime("%H:%M")
        except ValueError:
            continue
    raise ValueError(f"Cannot parse time: {s!r}")


def _parse_weeks(s: str) -> list[int]:
    """
    Extracts week numbers 1-5 from strings like '1-2 Week', '3-4-5 week'.
    """
    nums = _re.findall(r'\d+', str(s))
    return [int(n) for n in nums if 1 <= int(n) <= 5]


def _parse_day_sequence(s: str) -> list[str]:
    """
    Splits 'Mon, Tue, Wed' on ', ' and filters to valid day names only.
    """
    parts = str(s).split(", ")
    return [p.strip() for p in parts if p.strip() in _VALID_DAYS]


def _parse_date(s: str) -> date | None:
    """
    Tries formats: "%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%y".
    Returns None if all fail.
    """
    s = str(s).strip()
    for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_bulk_csv(df) -> tuple[list[dict], list[str]]:
    """
    Parses a pandas DataFrame (from CSV/Excel upload) into normalised config dicts.
    Returns (configs, warnings_list).
    """
    configs = []
    warnings_list = []

    for idx, row in df.iterrows():
        row_num = idx + 2  # 1-based, accounting for header

        try:
            course_name = str(row["Course Name"]).strip()
        except KeyError:
            warnings_list.append(f"Row {row_num}: Missing 'Course Name' column.")
            continue

        try:
            hours_per_day = int(row["Hours Per Day"])
        except (KeyError, ValueError, TypeError):
            warnings_list.append(f"Row {row_num} ({course_name}): Invalid or missing 'Hours Per Day'.")
            continue

        if hours_per_day <= 1:
            warnings_list.append(
                f"Row {row_num} ({course_name}): Hours Per Day={hours_per_day} is <= 1; skipping."
            )
            continue

        # Support both the typo column name and the correct one
        try:
            if "Toral Training Duration" in row.index:
                total_duration = int(row["Toral Training Duration"])
            elif "Total Training Duration" in row.index:
                total_duration = int(row["Total Training Duration"])
            else:
                total_duration = 24
        except (ValueError, TypeError):
            total_duration = 24

        training_days = total_duration // hours_per_day

        try:
            weeks = _parse_weeks(row["Weeks"])
        except KeyError:
            warnings_list.append(f"Row {row_num} ({course_name}): Missing 'Weeks' column.")
            continue
        if not weeks:
            warnings_list.append(
                f"Row {row_num} ({course_name}): No valid week numbers found in 'Weeks'={row.get('Weeks', '')}; skipping."
            )
            continue

        try:
            batch_type = str(row["Batch Type"]).strip().lower()
        except KeyError:
            batch_type = "weekday"

        try:
            day_sequence = _parse_day_sequence(row["Schedule Details"])
        except KeyError:
            warnings_list.append(f"Row {row_num} ({course_name}): Missing 'Schedule Details' column.")
            continue
        if not day_sequence:
            warnings_list.append(
                f"Row {row_num} ({course_name}): No valid days found in 'Schedule Details'={row.get('Schedule Details', '')}; skipping."
            )
            continue

        if len(day_sequence) < training_days:
            warnings_list.append(
                f"Row {row_num} ({course_name}): day_sequence has {len(day_sequence)} days but training_days={training_days}; using available days."
            )
        # Truncate to training_days
        day_sequence = day_sequence[:training_days]

        try:
            start_time = _parse_time_12h(row["Start Time"])
        except (KeyError, ValueError) as e:
            warnings_list.append(f"Row {row_num} ({course_name}): Bad Start Time — {e}; skipping.")
            continue

        try:
            end_time = _parse_time_12h(row["End Time"])
        except (KeyError, ValueError) as e:
            warnings_list.append(f"Row {row_num} ({course_name}): Bad End Time — {e}; skipping.")
            continue

        try:
            from_date = _parse_date(row["Start Date"])
        except KeyError:
            from_date = None
        if from_date is None:
            warnings_list.append(
                f"Row {row_num} ({course_name}): Cannot parse Start Date={row.get('Start Date', '')}; skipping."
            )
            continue

        try:
            to_date = _parse_date(row["End Date"])
        except KeyError:
            to_date = None
        if to_date is None:
            warnings_list.append(
                f"Row {row_num} ({course_name}): Cannot parse End Date={row.get('End Date', '')}; skipping."
            )
            continue

        anchor_us_dow = DAY_TO_US_DOW[day_sequence[0]]

        weeks_raw = str(row.get("Weeks", ""))
        schedule_raw = str(row.get("Schedule Details", ""))

        configs.append({
            "course_name":    course_name,
            "hours_per_day":  hours_per_day,
            "total_duration": total_duration,
            "training_days":  training_days,
            "weeks":          weeks,
            "batch_type":     batch_type,
            "day_sequence":   day_sequence,
            "start_time":     start_time,
            "end_time":       end_time,
            "from_date":      from_date,
            "to_date":        to_date,
            "anchor_us_dow":  anchor_us_dow,
            "weeks_raw":      weeks_raw,
            "schedule_raw":   schedule_raw,
        })

    return configs, warnings_list


def generate_schedules_bulk(csv_configs: list[dict], shared_params: dict) -> list[dict]:
    """
    Generates all schedule rows from bulk CSV configs.

    shared_params keys:
        pricing_tiers, default_capacity, training_mode, status,
        countries, usd_us_countries, course_id_map {name→id},
        override_from_date (date|None), override_to_date (date|None)
    """
    pricing_tiers    = shared_params["pricing_tiers"]
    default_capacity = shared_params["default_capacity"]
    training_mode    = shared_params["training_mode"]
    status           = shared_params["status"]
    countries        = shared_params["countries"]
    usd_us_countries  = set(shared_params.get("usd_us_countries", []))
    fixed_time_ranges = set(tuple(x) for x in shared_params.get("fixed_time_ranges", []))
    course_id_map     = shared_params.get("course_id_map", {})
    override_from        = shared_params.get("override_from_date")
    override_to          = shared_params.get("override_to_date")

    # Deduplicate configs by (course_name, hours_per_day, tuple(day_sequence), tuple(weeks), from_date, to_date)
    seen = set()
    unique_configs = []
    for cfg in csv_configs:
        key = (
            cfg["course_name"],
            cfg["hours_per_day"],
            tuple(cfg["day_sequence"]),
            tuple(cfg["weeks"]),
            cfg["from_date"],
            cfg["to_date"],
        )
        if key not in seen:
            seen.add(key)
            unique_configs.append(cfg)

    rows: list[dict] = []

    for cfg in unique_configs:
        course_name   = cfg["course_name"]
        course_id     = course_id_map.get(course_name, 0)
        hours_per_day = cfg["hours_per_day"]
        day_sequence  = cfg["day_sequence"]
        weeks         = cfg["weeks"]
        anchor_us_dow = cfg["anchor_us_dow"]
        start_time    = cfg["start_time"]
        end_time      = cfg["end_time"]
        batch_type    = cfg["batch_type"]

        from_date = override_from if override_from is not None else cfg["from_date"]
        to_date   = override_to   if override_to   is not None else cfg["to_date"]

        batches = _batch_dates_from_sequence(
            anchor_us_dow, day_sequence, weeks, from_date, to_date
        )

        if not batches:
            continue

        # Group batches by month
        by_month: dict[tuple, list] = {}
        for b in batches:
            key = (b[0].year, b[0].month)
            by_month.setdefault(key, []).append(b)

        # Sort months descending
        sorted_months = sorted(by_month.keys(), reverse=True)

        for country in countries:
            base_price = COURSE_PRICING.get(
                (course_id, country["region"]), 995
            )

            use_usd_us    = country["name"] in usd_us_countries
            eff_timezone  = _INPUT_TZ if use_usd_us else country["timezone"]
            eff_currency  = "USD"     if use_usd_us else country["currency"]
            eff_exch_rate = 1.0       if use_usd_us else country["exchange_rate"]

            for month_key in sorted_months:
                month_batches = sorted(by_month[month_key], key=lambda b: b[0], reverse=True)

                for start, end, sessions in month_batches:
                    sessions_json = json.dumps(
                        [s.strftime("%Y-%m-%d") for s in sessions]
                    )
                    use_fix_range   = (start_time, end_time) in fixed_time_ranges
                    if use_usd_us or use_fix_range:
                        _local_start_24 = start_time
                        _local_end_24   = end_time
                    else:
                        _local_start_24 = _convert_time(start_time, eff_timezone, start)
                        _local_end_24   = _convert_time(end_time,   eff_timezone, start)
                    local_start_time = to_12h(_local_start_24)
                    local_end_time   = to_12h(_local_end_24)
                    time_cat         = _time_category(_local_start_24)

                    for tier in pricing_tiers:
                        final_price = _calc_final_price(
                            base_price, tier["percentage"], eff_exch_rate
                        )
                        rows.append({
                            "Course ID":            course_id,
                            "Course Name":          course_name,
                            "Country ID":           country["id"],
                            "Country":              country["name"],
                            "Region":               country["region"],
                            "Pricing Tier ID":      tier["id"],
                            "Pricing Tier":         tier["name"],
                            "Duration (hr)":        hours_per_day,
                            "Batch Type":           batch_type,
                            "Start Date":           start.strftime("%Y-%m-%d"),
                            "End Date":             end.strftime("%Y-%m-%d"),
                            "Time Category":        time_cat,
                            "Session Dates (JSON)": sessions_json,
                            "Start Time":           local_start_time,
                            "End Time":             local_end_time,
                            "Timezone":             _tz_abbr(eff_timezone),
                            "Capacity":             default_capacity,
                            "Base Price USD":        base_price,
                            "Tier %":               tier["percentage"],
                            "Exchange Rate":         eff_exch_rate,
                            "Final Price":           final_price,
                            "Currency":              eff_currency,
                            "Training Mode":         training_mode,
                            "Status":                status.lower(),
                            "Generation Result":     "new",
                            "Error Message":         None,
                            "Price Override Flag":   "false",
                            "Override Reason":       None,
                        })

    return rows


def generate_schedules(params: dict) -> list[dict]:
    """
    Generate schedule rows from form params.

    Expected keys in `params`:
        course_id, course_name,
        from_date (date), to_date (date),
        pricing_tiers (list of tier dicts),
        training_days (int), default_capacity (int),
        weekday_batches_enabled (bool),
        weekday_day_format (list[str]), weekday_weeks (list[int]),
        weekend_batches_enabled (bool),
        weekend_day_format (list[str]), weekend_weeks (list[int]),
        training_mode (str), start_time (str), end_time (str),
        duration (int), status (str),
        countries (list of country dicts),

    Returns list of row dicts matching Excel column names.
    """
    weekday_batches = (
        _batch_dates_for_type(
            "weekday",
            params["weekday_day_format"],
            params["weekday_weeks"],
            params["from_date"],
            params["to_date"],
            params["training_days"],
        )
        if params.get("weekday_batches_enabled")
        else []
    )

    weekend_batches = (
        _batch_dates_for_type(
            "weekend",
            params["weekend_day_format"],
            params["weekend_weeks"],
            params["from_date"],
            params["to_date"],
            params["training_days"],
        )
        if params.get("weekend_batches_enabled")
        else []
    )

    # Group batches by month so we can interleave weekday then weekend per month
    def _by_month(batches):
        months: dict[tuple, list] = {}
        for b in batches:
            key = (b[0].year, b[0].month)
            months.setdefault(key, []).append(b)
        return months

    wd_by_month = _by_month(weekday_batches)
    we_by_month = _by_month(weekend_batches)
    all_months = sorted(set(list(wd_by_month) + list(we_by_month)), reverse=True)

    rows: list[dict] = []

    # Countries that keep USD pricing and New York timezone
    usd_us_countries: set[str] = set(params.get("usd_us_countries", []))

    for country in params["countries"]:
        base_price = COURSE_PRICING.get(
            (params["course_id"], country["region"]), 995
        )

        # Apply USD / US-timezone override if requested
        use_usd_us = country["name"] in usd_us_countries
        eff_timezone    = _INPUT_TZ              if use_usd_us else country["timezone"]
        eff_currency    = "USD"                  if use_usd_us else country["currency"]
        eff_exch_rate   = 1.0                    if use_usd_us else country["exchange_rate"]

        for month_key in all_months:
            for batch_type, batch_group in [
                ("weekday", wd_by_month.get(month_key, [])),
                ("weekend", we_by_month.get(month_key, [])),
            ]:
                for start, end, sessions in sorted(batch_group, key=lambda b: b[0], reverse=True):
                    sessions_json = json.dumps(
                        [s.strftime("%Y-%m-%d") for s in sessions]
                    )
                    # Convert times using the batch start date for correct DST
                    _local_start_24  = _convert_time(params["start_time"], eff_timezone, start)
                    _local_end_24    = _convert_time(params["end_time"],   eff_timezone, start)
                    local_start_time = to_12h(_local_start_24)
                    local_end_time   = to_12h(_local_end_24)
                    time_cat         = _time_category(_local_start_24)

                    for tier in params["pricing_tiers"]:
                        final_price = _calc_final_price(
                            base_price, tier["percentage"], eff_exch_rate
                        )
                        rows.append({
                            "Course ID":            params["course_id"],
                            "Course Name":          params["course_name"],
                            "Country ID":           country["id"],
                            "Country":              country["name"],
                            "Region":               country["region"],
                            "Pricing Tier ID":      tier["id"],
                            "Pricing Tier":         tier["name"],
                            "Duration (hr)":        params["duration"],
                            "Batch Type":           batch_type,
                            "Start Date":           start.strftime("%Y-%m-%d"),
                            "End Date":             end.strftime("%Y-%m-%d"),
                            "Time Category":        time_cat,
                            "Session Dates (JSON)": sessions_json,
                            "Start Time":           local_start_time,
                            "End Time":             local_end_time,
                            "Timezone":             _tz_abbr(eff_timezone),
                            "Capacity":             params["default_capacity"],
                            "Base Price USD":        base_price,
                            "Tier %":               tier["percentage"],
                            "Exchange Rate":         eff_exch_rate,
                            "Final Price":           final_price,
                            "Currency":              eff_currency,
                            "Training Mode":         params["training_mode"],
                            "Status":                params["status"].lower(),
                            "Generation Result":     "new",
                            "Error Message":         None,
                            "Price Override Flag":   "false",
                            "Override Reason":       None,
                        })

    return rows


def compute_tz_preview(
    countries: list[dict],
    start_time: str,
    end_time: str,
    ref_date: date,
    usd_us_countries,
    fixed_time_ranges=None,
) -> list[dict]:
    """
    Build a timezone-conversion preview table for the given countries.

    Mode per country:
      "Converted"  — standard DST-aware conversion from EST
      "Fixed Time" — keep original EST clock time, just relabel to local TZ
      "USD+Fixed"  — USD pricing + original EST clock time
    End Time is suffixed with " +1d" when it crosses midnight.
    """
    usd_set       = set(usd_us_countries)
    fix_set       = set(tuple(x) for x in (fixed_time_ranges or []))
    use_fix_range = (start_time, end_time) in fix_set
    rows = []
    for country in countries:
        use_usd_us = country["name"] in usd_set
        eff_tz = _INPUT_TZ if use_usd_us else country["timezone"]

        if use_usd_us or use_fix_range:
            local_start_24 = start_time
            local_end_24   = end_time
        else:
            local_start_24 = _convert_time(start_time, eff_tz, ref_date)
            local_end_24   = _convert_time(end_time,   eff_tz, ref_date)

        local_start = to_12h(local_start_24)
        local_end   = to_12h(local_end_24)

        sh, sm = map(int, local_start_24.split(":"))
        eh, em = map(int, local_end_24.split(":"))
        if (eh * 60 + em) < (sh * 60 + sm):
            local_end += " +1d"

        dt = datetime(ref_date.year, ref_date.month, ref_date.day, 12, 0,
                      tzinfo=ZoneInfo(eff_tz))
        offset_sec = int(dt.utcoffset().total_seconds())
        sign = "+" if offset_sec >= 0 else "-"
        abs_sec = abs(offset_sec)
        h_off, m_off = divmod(abs_sec // 60, 60)
        offset_str = f"UTC{sign}{h_off}:{m_off:02d}" if m_off else f"UTC{sign}{h_off}"

        if use_usd_us:
            mode = "USD + Fixed"
        elif use_fix_range:
            mode = "Fixed Time"
        else:
            mode = "Converted"

        rows.append({
            "Country":          country["name"],
            "Timezone":         _tz_abbr(eff_tz),
            "UTC Offset":       offset_str,
            "Local Start Time": local_start,
            "Local End Time":   local_end,
            "Time Mode":        mode,
        })
    return rows


def rows_to_excel_bytes(rows: list[dict]) -> bytes:
    """Serialize schedule rows to xlsx bytes. Includes Time Category column."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    COLUMNS = [
        "Course ID", "Course Name", "Country ID", "Country", "Region",
        "Pricing Tier ID", "Pricing Tier", "Duration (hr)", "Batch Type",
        "Start Date", "End Date", "Session Dates (JSON)", "Start Time",
        "End Time", "Timezone", "Capacity", "Base Price USD", "Tier %",
        "Exchange Rate", "Final Price", "Currency", "Training Mode",
        "Status", "Generation Result", "Error Message",
        "Price Override Flag", "Override Reason", "Time Category",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
